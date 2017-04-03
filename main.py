# -*- coding: utf-8 -*-


import openpyxl
import re
import nltk.classify

#import pandas

PathTrainExcelData=".\\data\\grp6_TrainedData2.xlsx"
PathExcludeWords=".\\data\\excludeWords.txt"
PathCleansedData=".\\data\\grp6_cleansed_data.xlsx"
PathAnalysedData=".\\data\\grp_analysedData.xlsx"

tweets=[]
meaningfulWordsList=[]
rowFeatureList=[]
meaningfulWordsList=[]


wb=openpyxl.load_workbook(PathTrainExcelData)
sheet=wb.get_sheet_by_name('Sheet1')
print "yoo"
print(sheet.max_row)
print "yoo2"

maxRows=sheet.max_row


def getExcludeWordsList(stopWordListFilePath):
    excludeWords=[]
    fp=open(stopWordListFilePath,'r')
    line=fp.readline()
    
    while line:
        word=line.strip()
        excludeWords.append(word)
        line=fp.readline()
    fp.close()
    return excludeWords

def replaceRepititiveChar(s):
    #look for 2 or more repetitions of character and replace with the character itself
    pattern = re.compile(r"(.)\1{1,}", re.DOTALL)
    return pattern.sub(r"\1\1", s)

def getFeatureList(tweet,excludeWords):
    featureList = []
    #split tweet into words
    #if tweet is not None:
    words = tweet.split()
    for w in words:
        #replace two or more with two occurrences
        w = replaceRepititiveChar(w)
        #strip punctuation
        w = w.strip('\'"?,.')
        #check if the word stats with an alphabet
        val = re.search(r"^[a-zA-Z][a-zA-Z0-9]*$", w)
        #ignore if it is a stop word
        if(len(w)>=3):
            if(w in excludeWords or val is None):
                continue
            else:
                featureList.append(w.lower())
        #for loop ends
    return featureList

def extract_features(tweetFeatureWords):
    tweet_words = set(tweetFeatureWords)
   # print "\nextract_features(demo tweet) \n"
   # print tweet_words
    #print "\n"
    featuresMap = {}
    for word in meaningfulWordsList:
        featuresMap['contains(%s)' % word] = (word in tweet_words)
#format is
# 'contains(abc)' : False
# 'contains(xyz)' : True
    #print featuresMap
    return featuresMap


excludeWords=getExcludeWordsList(PathExcludeWords)
#print excludeWords

for row in range(1,sheet.max_row+1):
    tweet=sheet['A'+str(row)].value
    sentiment=sheet['B'+str(row)].value
    rowFeatureList=getFeatureList(tweet,excludeWords)#only few meaningful words are selected (excludes excluding words list->precompiled)
    meaningfulWordsList.extend(rowFeatureList)
    tweets.append((rowFeatureList,sentiment))#above features list is appended with its sentiment(excel sheet)
    
meaningfulWordsList=list(set(meaningfulWordsList))#remove duplicates via set and this reconvert to list

training_set = nltk.classify.util.apply_features(extract_features,tweets,True)#3rd perimeter is false by default,therefore
#tweets(featurelist,sentiment)...featurelist will be passed to extract_features

       
print "\nprinting training set\n"
print training_set
print "\nyoo2\n"
#testTweet = 'conspiracy attack attacking'
#print extract_features(getFeatureList(testTweet,excludeWords))
#print getFeatureList(testTweet,excludeWords)

NBClassifier = nltk.NaiveBayesClassifier.train(training_set)

cleansedWorkBook=openpyxl.load_workbook(PathCleansedData)
cleansedDataExcelSheet=cleansedWorkBook.get_sheet_by_name('Sheet1')

#print cleansedDataExcelSheet.max_row
wb2=openpyxl.Workbook()
#wb2.save(PathAnalysedData)
#newSheet=wb.get_active_sheet()

#AnalysedWorkBook=openpyxl.load_workbook(PathAnalysedData)
AnalysedDataExcelSheet=wb2.active
#negColor =openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
#posColor=openpyxl.styles.Font(color=openpyxl.styles.colors.GREEN)
#neuColor=openpyxl.styles.Font(color=openpyxl.styles.colors.YELLOW)

for row in range(1,cleansedDataExcelSheet.max_row+1):
    testTweet=cleansedDataExcelSheet['A'+str(row)].value
    result=NBClassifier.classify(extract_features(getFeatureList(testTweet,excludeWords)))
    print str(row)+" " +testTweet + "\n"
    print result + "\n"
    AnalysedDataExcelSheet['A'+str(row)]=testTweet
    AnalysedDataExcelSheet['B'+str(row)]=result
    obj=AnalysedDataExcelSheet['B'+str(row)]
                   
    #NBClassifier.show_most_informative_features(10)

maxR=AnalysedDataExcelSheet.max_row
AnalysedDataExcelSheet.merge_cells('D3:G5')
AnalysedDataExcelSheet['D3']='TWEETS COUNT'
AnalysedDataExcelSheet['D6']="positive";
AnalysedDataExcelSheet['D7']="=COUNTIF(B1:B"+str(maxR)+",\"positive\")"       
AnalysedDataExcelSheet['E6']="negative";
AnalysedDataExcelSheet['E7']="=COUNTIF(B1:B"+str(maxR)+",\"negative\")"     
AnalysedDataExcelSheet['F6']="neutral";
AnalysedDataExcelSheet['F7']="=COUNTIF(B1:B"+str(maxR)+",\"neutral\")"   
AnalysedDataExcelSheet['G6']="Total"
AnalysedDataExcelSheet['G7']=maxR
wb2.save(PathAnalysedData)


print "finish"








